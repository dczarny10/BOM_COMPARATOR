import PySimpleGUI as sg
import openpyxl
import PLM_BOM_transpose
import os
from os import path as os_path
from os import environ

import mmc

sg.theme("DarkTeal12")

def popup_loading():
    sg.theme('DarkGrey')
    layout = [[sg.Text('Loading...',)]]
    window = sg.Window('Message', layout, no_titlebar=True, keep_on_top=True, finalize=True)
    sg.theme("DarkTeal12")
    return window

def PopupQuickView(subassembly):
    subassembly_revs = [i for i in data_subassemblies.keys() if subassembly[0] in i]
    if not subassembly_revs:
        sg.Popup("No such subassembly loaded")
        return None
    elif len(subassembly_revs) > 1:
        sg.Popup("More than one subassembly with this number loaded")
        return None

    subassembly_parts = sorted(data_subassemblies[subassembly_revs[0]]) #create list of parts that are in picked subassembly
    #declaration for QuickView table
    window_q = sg.Window("QuickView",
                         [[sg.Table(values=subassembly_parts, headings=headings, max_col_width=700,
                                    auto_size_columns=True,
                                    display_row_numbers=False,
                                    cols_justification=['l', 'c', 'l'],
                                    def_col_width=50,
                                    right_click_selects=True,
                                    num_rows=20,
                                    key='-TABLE_QUICKVIEW-',
                                    selected_row_colors='red on yellow',
                                    enable_events=True,
                                    expand_x=True,
                                    expand_y=True,
                                    enable_click_events=True,  # Comment out to not enable header and other clicks
                                    tooltip=f'This is BOM for {subassembly[0]}',
                                    right_click_menu=['', ['Quick view']])], ],
                         finalize=True, font=('Helvetica', 16),
                         resizable=True, size=(600, 200))
    while True:
        event_q, values_q = window_q.read()
        print(event_q, values_q)
        if event_q == sg.WIN_CLOSED:
            break
        if event_q == 'Quick view': #event for creating QuickView in QuickView
            window_q.close() #close primary window
            if not PopupQuickView(window_q['-TABLE_QUICKVIEW-'].get()[values_q['-TABLE_QUICKVIEW-'][0]]): #return None if picked part doesnt have subassembly
                PopupQuickView(subassembly) #create primary window if picked part doesnt have subassembly
    return True

def compare(source):
    parts_right, parts_left = [], []

    if len(values['-IN_RIGHT-']) > 0: #check if input is not empty
        if source == "input": #distinguish sources if writen by user or picked from BOX
            picked_right = values['-IN_RIGHT-'].replace("'", "").replace(",", "").replace("(", "").replace(")", "") #removes ',() from input
        else:
            picked_right = values['-BOX_RIGHT-'][0]
        if picked_right in data.keys():
            parts_right = sorted(data[picked_right])
            window['-TABLE_RIGHT-'].update(parts_right) #update Table
        else:
            window['-TABLE_RIGHT-'].update([])  # if user erased input set table to empty

    else:
        window['-TABLE_RIGHT-'].update([]) #if user erased input set table to empty

    if len(values['-IN_LEFT-']) > 0: #check if input is not empty
        if source == "input": #distinguish sources if writen by user or picked from BOX
            picked_left = values['-IN_LEFT-'].replace("'", "").replace(",", "").replace("(", "").replace(")", "") #removes ',() from input
        else:
            picked_left = values['-BOX_LEFT-'][0]
        if picked_left in data.keys():
            parts_left = sorted(data[picked_left])
            window['-TABLE_LEFT-'].update(parts_left) #update Table
        else:
            window['-TABLE_LEFT-'].update([])  # if user erased input set table to empty

    else:
        window['-TABLE_LEFT-'].update([]) #if user erased input set table to empty

    if len(window['-TABLE_RIGHT-'].get()) > 1 and len(window['-TABLE_LEFT-'].get()) > 1: #if both inputs are not empty
        for count, part in enumerate(parts_left): #enumerate through parts from left table
            index = -1
            for i, obj in enumerate(parts_right): #enumerate through parts from right table
                if obj[0] == part[0]: #if part found in right table
                    if obj[1] == part[1]: #if it have same quantity
                        index = i
                        break
                    index = -2
                    break
            if index == -1:
                window['-TABLE_LEFT-'].update(row_colors=((count, 'red'),)) #if part if not found in right table set row color to red
            elif index == -2:
                window['-TABLE_LEFT-'].update(row_colors=((count, 'black', 'yellow'),)) #if part if found in right table but with different quantity set row color to yellow

        for count, part in enumerate(parts_right): #enumerate through parts from right table
            index = -1
            for i, obj in enumerate(parts_left): #enumerate through parts from right table
                if obj[0] == part[0]: #if part found in left table
                    if obj[1] == part[1]: #if it have same quantity
                        index = i
                        break
                    index = -2
                    break
            if index == -1:
                window['-TABLE_RIGHT-'].update(row_colors=((count, 'red'),)) #if part if not found in right table set row color to red
            elif index == -2:
                window['-TABLE_RIGHT-'].update(row_colors=((count, 'black', 'yellow'),)) #if part if found in right table but with different quantity set row color to yellow
    else: #ensure no rows are colored if both inputs are not picked
        window['-TABLE_LEFT-'].update(row_colors=((0, ''),))
        window['-TABLE_RIGHT-'].update(row_colors=((0, ''),))

def sum_parts():
    for k in data.keys():  # iterate through all products
        # for c, part in enumerate(data[k]):
        #     for h in range(c, len(data[k])):
        #         if
        for part in data[k]:
            counted = [i[1] for i in data[k] if i[0]==part[0]]
            if len(counted) > 1:
                data[k] = [i for i in data[k] if i[0]!=part[0]]
                data[k].extend([(part[0], str(sum(map(int, counted))), part[2], part[3])])
    return True

def expand(parts_to_expand, products_to_expand):

    if parts_to_expand == "all_subassemblies":
        if not data_subassemblies:
            sg.Popup("No subassemblies loaded")
            return None
        for o in data_subassemblies.keys():  # iterate through all subassembly products
            subassembly_parts = data_subassemblies[o]
            if o.endswith("_SAP") or o.endswith("_PDF") or o.endswith("_MMC"):  # strip subassembly product number of _SAP or plant and revision level
                o = o[:-4]
            else:
                o = o[:-10]
            for k in data.keys():  # iterate through all products
                index = -1
                for i, obj in enumerate(data[k]):  # enumerate through loaded parts
                    if obj[0] == o:  # if same part
                        index = i
                        break
                if index > -1:
                    del data[k][index]  # delete original part
                    #data[k].extend([(i[0] + "_m", i[1], i[2]) for i in subassembly_parts])
                    data[k].extend([(i[0], i[1], i[2], i[3]) for i in subassembly_parts])

    else: #if extend only one part for all products #parts_to_expand == part number
        subassembly_revs = [i for i in data_subassemblies.keys() if parts_to_expand[0] in i]
        if not subassembly_revs:
            sg.Popup("No such subassembly loaded")
            return None
        elif len(subassembly_revs) > 1:
            sg.Popup("More than one subassembly with this part number loaded")
            #add option to pick which one to use
            return None
        if products_to_expand == "all_products":
            subassembly_parts = data_subassemblies[subassembly_revs[0]]
            for k in data.keys(): #iterate through all products
                index = -1
                for i, obj in enumerate(data[k]):  # enumerate through loaded parts
                    if obj[0] == parts_to_expand[0]:  # if same part
                        index = i
                        break
                if index > -1:
                    del data[k][index]  # delete original part
                    #data[k].extend([(i[0] + "_m", i[1], i[2]) for i in subassembly_parts])  # add subassembly parts to parts list
                    data[k].extend([(i[0], i[1], i[2], i[3]) for i in subassembly_parts])  # add subassembly parts to parts list

        else:#if extend only one part for one product
            subassembly_parts = data_subassemblies[subassembly_revs[0]]
            products_to_expand = products_to_expand.replace("'", "").replace(",", "").replace("(", "").replace(")", "")
            index = -1
            for i, obj in enumerate(data[products_to_expand]):  # enumerate through loaded parts
                if obj[0] == parts_to_expand[0]:  # if same part
                    index = i
                    break
            del data[products_to_expand][index] #delete original part
           #data[products_to_expand].extend([(i[0]+"_m", i[1], i[2]) for i in subassembly_parts]) #add subassembly parts to parts list
            data[products_to_expand].extend([(i[0], i[1], i[2], i[3]) for i in subassembly_parts])  # add subassembly parts to parts list


    sum_parts()
    return True

def mass_check():
    if len(data) < 1:
        sg.Popup("No parts loaded")
        return False

    window_m = sg.Window("PLM vs SAP mass check",
                         [[sg.Text('Quickly check all differences between SAP and PLM')],
                          [sg.Text('Make sure both data from SAP and PLM are loaded')],
                          [sg.Checkbox('Expand all subassemblies:', default=False, key='-EXPAND-')],
                          [sg.Checkbox('Ignore documentation without quantity:', default=False, key='-DOC-')],
                          [sg.Text('Ignore parts with specific name, separate with ;'), sg.Input('', key='-IGNORE_NAME-')],
                          [sg.Input('', key='-SAVE-FILE-PATH-'), sg.Button('Save file', key='-SAVE-FILE-BUTTON-')],
                          [sg.Button('OK')], ],
                         finalize=True, font=('Helvetica', 16),
                         resizable=True, size=(750, 300))

    while True:
        event_m, values_m = window_m.read()
        #print(event_m, values_m)
        if event_m == sg.WIN_CLOSED:
            break
        elif event_m == '-SAVE-FILE-BUTTON-':
            save_file = sg.PopupGetFile('Select file', no_window=True, multiple_files=False, save_as= True, default_extension='xlsx',
                            file_types=(("Excel file", [".xlsx"]),))
            window_m['-SAVE-FILE-PATH-'].update(save_file)

        elif event_m == 'OK':
            if not values_m['-SAVE-FILE-PATH-']:
                sg.popup_error("No path for save file specified")
                continue

            if os.path.exists(values_m['-SAVE-FILE-PATH-']):
                try:
                    os.rename(values_m['-SAVE-FILE-PATH-'], values_m['-SAVE-FILE-PATH-'])
                except:
                    sg.popup_error("Please close specified file")
                    continue

            ignore = [s.strip() for s in values_m['-IGNORE_NAME-'].split(';')]
            compared = []
            products_pairs = []
            redFill = openpyxl.styles.PatternFill(start_color='FFFF0000',
                                  end_color='FFFF0000',
                                  fill_type='solid')
            yellowFill = openpyxl.styles.PatternFill(start_color='FFFF00',
                                  end_color='FFFF00',
                                  fill_type='solid')
            # header = ('Part name', 'Part number', 'Quantity PLM/PDF', 'Quanity SAP', 'OK?/NOK?')
            index_not_found = 7 + len(ignore)
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.column_dimensions['A'].width = 50
            ws1.title = "Info"
            ws1['A1'] = "Data generated by mass check"
            ws1['A2'] = "Each product have an individual tab with report"
            ws1['A4'] = "Parts that were omitted during analysis:"
            if values_m['-DOC-']:
                ws1['A5'] = 'all parts without quantity'
            for index, ignored in enumerate(ignore):
                ws1.cell(row=index+6, column=1).value = ignored
            ws1.cell(row=index_not_found, column=1).value = "Products where no pair was found:"
            if values_m['-EXPAND-']:
                expand("all_subassemblies", "all_products")
            products_list = sorted(tuple(set([i[:-4] if i[-4] == '_' else i[:-10] for i in data.keys()]))) #create a products lists without _SAP or revision

            for p in products_list:
                products_pairs.append(sorted([i for i in data.keys() if i.startswith(p)])) #create pairs of SAP and PLM product number
            for count, j in enumerate(products_pairs):
                if len(j) < 2: #no pair in loaded data
                    ws1.cell(row=index_not_found, column=1).value = j[0] #write product number
                    index_not_found += 1
                else:
                    header = ('Part name', 'Part number', 'Quantity PDF', 'Quantity SAP', 'OK?/NOK?')
                    ws = wb.create_sheet(title=products_list[count]) #create new sheet with product number as title
                    ws.column_dimensions['A'].width = 35
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 13
                    ws.column_dimensions['D'].width = 13
                    ws.column_dimensions['E'].width = 10
                    row_index = 2
                    for col, t in enumerate(header): #loop to write header
                        ws.cell(row=1, column=col+1).value = t
                    for part_PLM in data[j[0]]:
                        if values_m['-DOC-'] and part_PLM[1] in ('', ' ', 'None', None):
                            continue
                        if part_PLM[2] in ignore: #ignore parts with specified names
                            continue
                        index = [-1]
                        for part_SAP in data[j[1]]:
                            if part_PLM[0] == part_SAP[0]: #if part found
                                if part_PLM[1] == part_SAP[1]: #if same quantity
                                    index = [1, part_SAP[1]]
                                    break
                                else: #if not same quanity
                                    index = [0, part_SAP[1]]
                                    break
                        if index[0] == 1:
                            #compared.append((part_PLM[2], part_PLM[0], part_PLM[1], index[1], "OK"))
                            ws.cell(row=row_index, column=1).value = part_PLM[2]
                            ws.cell(row=row_index, column=2).value = part_PLM[0]
                            ws.cell(row=row_index, column=3).value = part_PLM[1]
                            ws.cell(row=row_index, column=4).value = index[1]
                            ws.cell(row=row_index, column=5).value = "OK"
                        elif index[0] == 0:
                            #compared.append((part_PLM[2], part_PLM[0], part_PLM[1], index[1], "NOK"))
                            ws.cell(row=row_index, column=1).value = part_PLM[2]
                            ws.cell(row=row_index, column=2).value = part_PLM[0]
                            ws.cell(row=row_index, column=3).value = part_PLM[1]
                            ws.cell(row=row_index, column=4).value = index[1]
                            ws.cell(row=row_index, column=5).value = "NOK"
                            ws.cell(row=row_index, column=5).fill = yellowFill
                            ws.sheet_properties.tabColor = 'FFFF00'
                        else:
                            #compared.append((part_PLM[2], part_PLM[0], part_PLM[1], "0", "NOK"))
                            ws.cell(row=row_index, column=1).value = part_PLM[2]
                            ws.cell(row=row_index, column=2).value = part_PLM[0]
                            ws.cell(row=row_index, column=3).value = part_PLM[1]
                            ws.cell(row=row_index, column=4).value = "0"
                            ws.cell(row=row_index, column=5).value = "NOK"
                            ws.cell(row=row_index, column=5).fill = redFill
                            ws.sheet_properties.tabColor = 'FFFF0000'


                        row_index += 1

                    #for row, row_text in enumerate(compared): #loop to write data
                        #for col, t in enumerate(row_text):
                            #ws.cell(row=row+2, column=col+1).value = t

                    for part_SAP in data[j[1]]:
                        index = -1
                        if values_m['-DOC-'] and part_SAP[1] in ('', ' '):
                            continue
                        if part_SAP[2] in ignore: #ignore parts with specified names
                            continue
                        for part_PLM in data[j[0]]:
                            if part_PLM[0] == part_SAP[0]:  # if part found
                                index = 1
                                break
                        if index == -1:
                            ws.cell(row=row_index, column=1).value = part_SAP[2]
                            ws.cell(row=row_index, column=2).value = part_SAP[0]
                            ws.cell(row=row_index, column=3).value = "0"
                            ws.cell(row=row_index, column=4).value = part_SAP[1]
                            ws.cell(row=row_index, column=5).value = "NOK"
                            ws.cell(row=row_index, column=5).fill = redFill
                            ws.sheet_properties.tabColor = 'FFFF0000'
                            row_index += 1

            wb.save(filename = values_m['-SAVE-FILE-PATH-'])
            sg.Popup("Successful")
            window_m.close()
            break
    return True

def mmc_check():
    window.hide()
    window.disable()
    menu_def_mmc = [['&File', ['&Exit']],
                ['&Edit', [], ],
                ['&Tools', ['&Compare BOMs', '&SAP/PLM Mass Check', '&MMC Mass Encode'], ],
                ['&Help', '&About...'], ]
    layout_mmc =[[sg.Menu(menu_def_mmc, )],
                 [sg.Text('Cincom logic file')],
                 [sg.Input('', key='-CINCOM-FILE-PATH-'), sg.Button('Pick file', key='-CINCOM-FILE-BUTTON-')],
                 [sg.Text('List of MMC files to encode. 1st column is part name, 2nd MMC')],
                 [sg.Input('', key='-MMC-FILE-PATH-'), sg.Button('Pick file', key='-MMC-FILE-BUTTON-')],
                 [sg.Text('Target file with encoded BOMs')],
                 [sg.Input('', key='-BOM-FILE-PATH-'), sg.Button('Pick file', key='-BOM-FILE-BUTTON-')],
                 [sg.Text('Exclude specific tabs (need full name, separate with ;)')],
                 [sg.Input('', key='-MMC-EXCLUDE-')],
                 [sg.HSeparator()],
                 [sg.Text('Progress: '),sg.Text('', key='-MMC_PROGRESS-')],
                 [sg.VPush()],
                 [sg.Push(), sg.Button('OK')],

                 ]

    window_mmc = sg.Window('MMC check', layout_mmc,
                           return_keyboard_events=True, finalize=True, font=('Helvetica', 16),
                           resizable=True, size=(1000, 700))
    window_mmc.move(window.current_location(more_accurate = False, without_titlebar = False)[0], window.current_location(more_accurate = False, without_titlebar = False)[1])

    while True:
        event_mmc, values_mmc = window_mmc.read()
        #print(event_mmc, values_mmc)
        if event_mmc in (sg.WIN_CLOSED, 'Exit'):
            window.close()
            break
        elif event_mmc == 'Compare BOMs':
            window.move(window_mmc.current_location(more_accurate=False, without_titlebar=False)[0],
                        window_mmc.current_location(more_accurate=False, without_titlebar=False)[1])
            window.un_hide()
            window.enable()
            break
        elif event_mmc == 'SAP/PLM Mass Check':
            mass_check()
        elif event_mmc == '-CINCOM-FILE-BUTTON-':
            save_file = sg.PopupGetFile('Select file', no_window=True, multiple_files=False, save_as= False, default_extension='xlsx',
                            file_types=(("Excel file", [".xlsx"]),))
            window_mmc['-CINCOM-FILE-PATH-'].update(save_file)

        elif event_mmc == '-MMC-FILE-BUTTON-':
            save_file = sg.PopupGetFile('Select file', no_window=True, multiple_files=False, save_as= False, default_extension='xlsx',
                            file_types=(("Excel file", [".xlsx"]),))
            window_mmc['-MMC-FILE-PATH-'].update(save_file)

        elif event_mmc == '-BOM-FILE-BUTTON-':
            save_file = sg.PopupGetFile('Select file', no_window=True, multiple_files=False, save_as= True, default_extension='xlsx',
                            file_types=(("Excel file", [".xlsx"]),))
            window_mmc['-BOM-FILE-PATH-'].update(save_file)



        elif event_mmc == 'OK':
            if not values_mmc['-CINCOM-FILE-PATH-']:
                sg.popup_error("No path for CINCOM logic file specified")
                continue
            if not values_mmc['-MMC-FILE-PATH-']:
                sg.popup_error("No path for MMC file specified")
                continue
            if not values_mmc['-BOM-FILE-PATH-']:
                sg.popup_error("No path for target file specified")
                continue
            window_mmc.perform_long_operation(lambda: mmc.encode_mmc(values_mmc['-MMC-FILE-PATH-'], values_mmc['-CINCOM-FILE-PATH-'], values_mmc['-BOM-FILE-PATH-'], values_mmc['-MMC-EXCLUDE-'].split(';') , window_mmc), '-BOM_DONE-')

        elif event_mmc == '-BOM_DONE-':
            if values_mmc['-BOM_DONE-'] == 1:
                sg.popup('Completed')
                window_mmc['-MMC_PROGRESS-'].update('')
            elif values_mmc['-BOM_DONE-'] == 'permission_error':
                sg.popup_error('Permission denied\nPlease close target file')

    window_mmc.close()

def save_bom():
    if len(data) < 1:
        sg.Popup("No parts loaded")
        return False
    save_file = sg.PopupGetFile('Select file', no_window=True, multiple_files=False, save_as=True, default_extension='xlsx',
                                file_types=(("Excel file", [".xlsx"]),))
    if os.path.exists(save_file):
        try:
            os.rename(save_file, save_file)
        except:
            sg.popup_error("Please close specified file")
            return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Position', 'Part', 'Qty', 'Description', 'Parent'])
    for key in data:
        for row in data[key]:
            try:
                ws.append([row[3], row[0], row[1], row[2], key[:-4] if key[-4] == '_' else key[:-10]])
            except:
                print(key)
                print(row)

    wb.save(filename=save_file)
    sg.Popup("Successful")



headings = ["Part", "Qty", "Description"] #heading for tables

menu_def = [['&File', ['&Load data', '&Load subassemblies', '&Clear data', '&Save data', '&Exit']],
                ['&Edit', ['&Expand all subassemblies' ],],
                ['&Tools', ['&Compare BOMs', '&SAP/PLM Mass Check', '&MMC Mass Encode'],],
                ['&Help', '&About...'], ]

menu_top = [['&Compare'],['&Mass check'], ]
#column declaration with left table
col1 = sg.Column([[sg.Table(values=[[]], headings=headings, max_col_width=700,
                            auto_size_columns=False,
                            display_row_numbers=False,
                            # justification='center',
                            cols_justification=['l', 'c', 'l'],
                            #def_col_width=50,
                            col_widths=[15, 5, 30],
                            right_click_selects=True,
                            num_rows=20,
                            key='-TABLE_LEFT-',
                            selected_row_colors='white on blue',
                            enable_events=True,
                            expand_x=True,
                            expand_y=True,
                            enable_click_events=True,  # Comment out to not enable header and other clicks
                            tooltip='This is BOM1',
                            right_click_menu=['', ['Quick view', 'Expand for this', 'Expand for all']])]],
                 pad=(0, 0), expand_x=True)

#column declaration with right table
col2 = sg.Column([[sg.Table(values=[[]], headings=headings, max_col_width=700,
                            auto_size_columns=False,
                            display_row_numbers=False,
                            # justification='right',
                            cols_justification=['l', 'c', 'l'],
                            #def_col_width=50,
                            col_widths=[10, 5, 30],
                            right_click_selects=True,
                            num_rows=20,
                            key='-TABLE_RIGHT-',
                            selected_row_colors='white on blue',
                            enable_events=True,
                            expand_x=True,
                            expand_y=True,
                            enable_click_events=True,  # Comment out to not enable header and other clicks
                            tooltip='This is BOM2',
                            right_click_menu=['', ['Quick view', 'Expand for this', 'Expand for all']])]],
                 pad=(0, 0), expand_x=True)

#column declaration with top buttons
col3 = sg.Column([[sg.Frame('Actions:',
                            [[sg.Column([[sg.Button('Clear Tables', key='-CLEAR_TABLES-')]],
                    pad=(0, 0))]])]], pad=(0, 0))

#column declaration with left input
col4 = sg.Column([
    [sg.Text('Product number:')],
    [sg.Input(size=(20, 1), enable_events=True, key='-IN_LEFT-')],
    [sg.pin(sg.Col([[sg.Listbox(values=[], size=(20, 4), enable_events=True, key='-BOX_LEFT-',
                                select_mode=sg.LISTBOX_SELECT_MODE_SINGLE, no_scrollbar=True)]],
                   key='-BOX-CONTAINER_LEFT-', pad=(0, 0), visible=False))]])

#column declaration with right input
col5 = sg.Column([
    [sg.Text('Product number:')],
    [sg.Input(size=(20, 1), enable_events=True, key='-IN_RIGHT-')],
    [sg.pin(sg.Col([[sg.Listbox(values=[], size=(20, 4), enable_events=True, key='-BOX_RIGHT-',
                                select_mode=sg.LISTBOX_SELECT_MODE_SINGLE, no_scrollbar=True)]],
                   key='-BOX-CONTAINER_RIGHT-', pad=(0, 0), visible=False))]])

layout = [[sg.Menu(menu_def, )],
          [col4, sg.Push(), col3, sg.Push(), col5],
          [col1, col2],]

window = sg.Window('Ultimate BOM Comparator', layout, return_keyboard_events=True, finalize=True, font=('Helvetica', 16),
                   resizable=True, size=(1500, 700))


list_element_left: sg.Listbox = window.Element(
    '-BOX_LEFT-')  # store listbox element for easier access and to get to docstrings
list_element_right: sg.Listbox = window.Element(
    '-BOX_RIGHT-')  # store listbox element for easier access and to get to docstrings
prediction_list_left, input_text_left, sel_item_left = [], "", 0
prediction_list_right, input_text_right, sel_item_right = [], "", 0
data = {}
data_subassemblies = {}

table_clicked_right, table_clicked_left = False, False

while True:
    event, values = window.read()
    print(event, values)
    if event in (sg.WIN_CLOSED,'Exit'):
        break

    elif event == 'Load data':
        try:
            files = sg.PopupGetFile('Select folder', no_window=True, multiple_files=True,
                                    file_types=(("Excel files", [".xls", ".xlsx"]),))
            popup_load = popup_loading()
            window.perform_long_operation(lambda: PLM_BOM_transpose.load_files(files), '-DATA-LOADING-DONE-')
        except Exception as e:
            print(e)
            sg.Popup(f'Wrong file\n{e}', keep_on_top=True)

    elif event == 'Load subassemblies':
        try:
            files = sg.PopupGetFile('Select folder', no_window=True, multiple_files=True,
                                    file_types=(("Excel files", [".xls", ".xlsx"]),))
            popup_load = popup_loading()
            window.perform_long_operation(lambda: PLM_BOM_transpose.load_files(files), '-SUBASSEMBLY-LOADING-DONE-')
        except Exception as e:
            print(e)
            sg.Popup(f'Wrong file\n{e}', keep_on_top=True)

    elif event == 'Save data':
        save_bom()

    elif event == '-SUBASSEMBLY-LOADING-DONE-':
        data_subassemblies.update(values['-SUBASSEMBLY-LOADING-DONE-'])
        popup_load.close()

    elif event == '-DATA-LOADING-DONE-':
        data.update(values['-DATA-LOADING-DONE-'])
        popup_load.close()


    elif event == "Quick view":
        if not table_clicked_right:
            try:
                PopupQuickView(window['-TABLE_LEFT-'].get()[table_clicked_left])
            except:
                continue
        elif not table_clicked_left:
            try:
                PopupQuickView(window['-TABLE_RIGHT-'].get()[table_clicked_right])
            except:
                continue

    elif event == "Expand for this":
        if not table_clicked_right:
            try:
                expand(window['-TABLE_LEFT-'].get()[table_clicked_left], window['-IN_LEFT-'].get())
                compare("input")
            except Exception as e:
                print(e)
                continue
        elif not table_clicked_left:
            try:
                expand(window['-TABLE_RIGHT-'].get()[table_clicked_right], window['-IN_RIGHT-'].get())
                compare("input")
            except:
                continue

    elif event == "Expand for all":
        if not table_clicked_right:
            try:
                expand(window['-TABLE_LEFT-'].get()[table_clicked_left], "all_products")
                compare("input")
            except Exception as e:
                print(e)
                continue
        elif not table_clicked_left:
            try:
                expand(window['-TABLE_RIGHT-'].get()[table_clicked_right], "all_products")
                compare("input")
            except:
                continue

    elif event == 'Expand all subassemblies':
        try:
            expand("all_subassemblies", "all_products")
            window['-TABLE_LEFT-'].update([])
            window['-TABLE_RIGHT-'].update([])
            window['-TABLE_LEFT-'].update(row_colors=((0, ''),))
            window['-TABLE_RIGHT-'].update(row_colors=((0, ''),))
            compare("input")
        except Exception as e:
            print(e)
            continue

    elif event == 'SAP/PLM Mass Check':
        mass_check()

    elif isinstance(event, tuple): #event to recognize which row in which table was clicked by user. Needed for rightclickmenu
        if event[0] == '-TABLE_LEFT-':
            table_clicked_left = event[2][0]
            table_clicked_right = False
        elif event[0] == '-TABLE_RIGHT-':
            table_clicked_right = event[2][0]
            table_clicked_left = False

    elif event == '-CLEAR_TABLES-':
        window['-IN_LEFT-'].update('')
        window['-BOX-CONTAINER_LEFT-'].update(visible=False)
        window['-IN_RIGHT-'].update('')
        window['-BOX-CONTAINER_RIGHT-'].update(visible=False)
        window['-TABLE_LEFT-'].update([])
        window['-TABLE_RIGHT-'].update([])

    elif event == 'Clear data':
        data, data_subassemblies = {}, {}
        window['-IN_LEFT-'].update('')
        window['-BOX-CONTAINER_LEFT-'].update(visible=False)
        window['-IN_RIGHT-'].update('')
        window['-BOX-CONTAINER_RIGHT-'].update(visible=False)
        window['-TABLE_LEFT-'].update([])
        window['-TABLE_RIGHT-'].update([])

    elif event == '\r':
        if len(values['-BOX_LEFT-']) > 0:
            window['-IN_LEFT-'].update(value=values['-BOX_LEFT-'])
            window['-BOX-CONTAINER_LEFT-'].update(visible=False)
        if len(values['-BOX_RIGHT-']) > 0:
            window['-IN_RIGHT-'].update(value=values['-BOX_RIGHT-'])
            window['-BOX-CONTAINER_RIGHT-'].update(visible=False)

    elif event == '-IN_LEFT-':
        text = values['-IN_LEFT-']
        if text == input_text_left:
            continue
        else:
            input_text_left = text
        prediction_list_left = []
        if text:
            prediction_list_left = [item for item in data.keys() if item.startswith(text)]

        list_element_left.update(values=prediction_list_left)
        sel_item_left = 0
        list_element_left.update(set_to_index=sel_item_left)

        if len(prediction_list_left) > 0:
            window['-BOX-CONTAINER_LEFT-'].update(visible=True)
        else:
            window['-BOX-CONTAINER_LEFT-'].update(visible=False)
        compare("input")

    elif event == '-IN_RIGHT-':
        text = values['-IN_RIGHT-']
        if text == input_text_right:
            continue
        else:
            input_text_right = text
        prediction_list_right = []
        if text:
            prediction_list_right = [item for item in data.keys() if item.startswith(text)]

        list_element_right.update(values=prediction_list_right)
        sel_item_right = 0
        list_element_right.update(set_to_index=sel_item_right)

        if len(prediction_list_right) > 0:
            window['-BOX-CONTAINER_RIGHT-'].update(visible=True)
        else:
            window['-BOX-CONTAINER_RIGHT-'].update(visible=False)

        compare("input")

    elif event == '-BOX_LEFT-':
        window['-IN_LEFT-'].update(value=values['-BOX_LEFT-'])
        window['-BOX-CONTAINER_LEFT-'].update(visible=False)
        compare("box")

    elif event == '-BOX_RIGHT-':
        window['-IN_RIGHT-'].update(value=values['-BOX_RIGHT-'])
        window['-BOX-CONTAINER_RIGHT-'].update(visible=False)
        compare("box")

    elif event == 'About...':
        sg.popup("Simple app created to make work easier :)\n"
                 "In case of any bugs/comments please contact me at:\nd.czarnecki@whitedriveproducts.com",
            title = "About...",
            keep_on_top = True,)

    elif event == 'MMC Mass Encode':
        mmc_check()

window.close()
