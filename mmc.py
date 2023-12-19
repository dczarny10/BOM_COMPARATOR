import pandas as pd
import numpy as np
import openpyxl

# adr = {'ADR_F_01': mmc[0:3],
# 'ADR_F_02': mmc[3:4],
# 'ADR_F_03': mmc[4:5],
# 'ADR_F_04': mmc[5:6],
# 'ADR_F_05': mmc[6:7],
# 'ADR_F_06': mmc[7:9],
# 'ADR_F_07': mmc[9:10],
# 'ADR_F_08': mmc[10:11],
# 'ADR_F_09': mmc[11:12],
# 'ADR_F_10': mmc[12:14],
# 'ADR_F_11': mmc[14:16],
# 'ADR_F_12': mmc[16:18],
# 'ADR_F_13': mmc[18:20],
# 'ADR_F_14': mmc[20:21],
# 'ADR_F_15': mmc[21:23],
# 'ADR_F_16': mmc[23:24],
# 'ADR_F_17': mmc[24:25],
# 'ADR_F_18': mmc[25:26],
# 'ADR_F_19': mmc[26:27],
# 'ADR_F_20': mmc[27:29],
# 'ADR_F_21': mmc[29:30],
# 'ADR_F_22': mmc[30:31],
# 'ADR_F_23': mmc[31:32]}

# acc = {'ACC_f_01': mmc[0:3],
# 'ACC_f_02': mmc[3:4],
# 'ACC_f_03': mmc[4:5],
# 'ACC_f_04': mmc[5:6],
# 'ACC_f_05': mmc[6:8],
# 'ACC_f_06': mmc[8:9],
# 'ACC_f_07': mmc[9:10],
# 'ACC_f_08': mmc[10:11],
# 'ACC_f_09': mmc[11:13],
# 'ACC_f_10': mmc[13:15],
# 'ACC_f_11': mmc[15:17],
# 'ACC_f_12': mmc[17:18],
# 'ACC_f_13': mmc[18:20],
# 'ACC_f_14': mmc[20:21],
# 'ACC_f_15': mmc[21:22],
# 'ACC_f_16': mmc[22:23],
# 'ACC_f_17': mmc[23:24],
# 'ACC_f_18': mmc[24:26],
# 'ACC_f_19': mmc[26:27],
# 'ACC_f_20': mmc[27:28],
# 'ACC_f_21': mmc[28:29]}
def encode_mmc(mmc_list, constrains, target_file, excluded, window):
    mmc_list_file = pd.read_excel(mmc_list)
    workbook = openpyxl.Workbook()
    sheet = workbook['Sheet']
    sheet.title = 'mmc'
    sheet.append(['Position', 'Part', 'Qty', 'Description', 'Parent', 'MMC'])
    sheet.column_dimensions['A'].width = 9
    sheet.column_dimensions['B'].width = 17
    sheet.column_dimensions['C'].width = 8
    sheet.column_dimensions['D'].width = 40
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 41

    # sheet_conflicts = workbook.create_sheet('Conflicting conditions')
    # sheet_conflicts.append((['Parent part', 'MMC', 'TAB', 'Possible parts']))
    # sheet_conflicts.append((['', '', '', 'Part A', 'Qty A', 'Part B', 'Qty B']))
    # sheet_conflicts.merge_cells('A1:A2')
    # sheet_conflicts.merge_cells('B1:B2')
    # sheet_conflicts.merge_cells('C1:C2')
    # sheet_conflicts.merge_cells('D1:G1')
    # sheet_conflicts.column_dimensions['A'].width = 16
    # sheet_conflicts.column_dimensions['B'].width = 41
    # sheet_conflicts.column_dimensions['C'].width = 32
    # sheet_conflicts.column_dimensions['D'].width = 17
    # sheet_conflicts.column_dimensions['E'].width = 10
    # sheet_conflicts.column_dimensions['F'].width = 17
    # sheet_conflicts.column_dimensions['G'].width = 10
    # sheet_conflicts.cell(1, 1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    # sheet_conflicts.cell(1, 2).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    # sheet_conflicts.cell(1, 3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    # sheet_conflicts.cell(1, 4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


    try:
        workbook.save(filename=target_file)
    except PermissionError:
        return 'permission_error'


    xls = pd.ExcelFile(constrains) #!!! Will only decode 1 family at a time !!!

    for i, row in mmc_list_file.iterrows():
        window['-MMC_PROGRESS-'].update(row[0])
        if row[1][0:3] == 'ADR': #S10
            mmc = {'ADR_F_01': row[1][0:3],
                   'ADR_F_02': row[1][3:4],
                   'ADR_F_03': row[1][4:5],
                   'ADR_F_04': row[1][5:6],
                   'ADR_F_05': row[1][6:7],
                   'ADR_F_06': row[1][7:9],
                   'ADR_F_07': row[1][9:10],
                   'ADR_F_08': row[1][10:11],
                   'ADR_F_09': row[1][11:12],
                   'ADR_F_10': row[1][12:14],
                   'ADR_F_11': row[1][14:16],
                   'ADR_F_12': row[1][16:18],
                   'ADR_F_13': row[1][18:20],
                   'ADR_F_14': row[1][20:21],
                   'ADR_F_15': row[1][21:23],
                   'ADR_F_16': row[1][23:24],
                   'ADR_F_17': row[1][24:25],
                   'ADR_F_18': row[1][25:26],
                   'ADR_F_19': row[1][26:27],
                   'ADR_F_20': row[1][27:29],
                   'ADR_F_21': row[1][29:30],
                   'ADR_F_22': row[1][30:31],
                   'ADR_F_23': row[1][31:32]}

        elif row[1][0:3] == 'ACC': #S20
            mmc = {'ACC_f_01': row[1][0:3],
                   'ACC_f_02': row[1][3:4],
                   'ACC_f_03': row[1][4:5],
                   'ACC_f_04': row[1][5:6],
                   'ACC_f_05': row[1][6:8],
                   'ACC_f_06': row[1][8:9],
                   'ACC_f_07': row[1][9:10],
                   'ACC_f_08': row[1][10:11],
                   'ACC_f_09': row[1][11:13],
                   'ACC_f_10': row[1][13:15],
                   'ACC_f_11': row[1][15:17],
                   'ACC_f_12': row[1][17:18],
                   'ACC_f_13': row[1][18:20],
                   'ACC_f_14': row[1][20:21],
                   'ACC_f_15': row[1][21:22],
                   'ACC_f_16': row[1][22:23],
                   'ACC_f_17': row[1][23:24],
                   'ACC_f_18': row[1][24:26],
                   'ACC_f_19': row[1][26:27],
                   'ACC_f_20': row[1][27:28],
                   'ACC_f_21': row[1][28:29]}

        elif row[1][0:3] == 'ACD': #VIS30
            mmc = {'V30_f_01': row[1][0:3],
                   'V30_f_02': row[1][3:5],
                   'V30_f_03': row[1][5:6],
                   'V30_f_04': row[1][6:8],
                   'V30_f_05': row[1][8:9],
                   'V30_f_06': row[1][9:10],
                   'V30_f_07': row[1][10:11],
                   'V30_f_08': row[1][11:13],
                   'V30_f_09': row[1][13:14],
                   'V30_f_10': row[1][14:15],
                   'V30_f_11': row[1][15:16]}

        elif row[1][0:3] == 'ADL': #VIS40
            mmc = {'V40_f_01': row[1][0:3],
                   'V40_f_02': row[1][3:5],
                   'V40_f_03': row[1][5:6],
                   'V40_f_04': row[1][6:8],
                   'V40_f_05': row[1][8:9],
                   'V40_f_06': row[1][9:10],
                   'V40_f_07': row[1][10:11],
                   'V40_f_08': row[1][11:13],
                   'V40_f_09': row[1][13:14],
                   'V40_f_10': row[1][14:15],
                   'V40_f_11': row[1][15:16]}

        elif row[1][0:3] == 'ACB': #VIS45
            mmc = {'V45_f_01': row[1][0:3],
                   'V45_f_02': row[1][3:5],
                   'V45_f_03': row[1][5:6],
                   'V45_f_04': row[1][6:8],
                   'V45_f_05': row[1][8:9],
                   'V45_f_06': row[1][9:10],
                   'V45_f_07': row[1][10:11],
                   'V45_f_08': row[1][11:13],
                   'V45_f_09': row[1][13:14],
                   'V45_f_10': row[1][14:15],
                   'V45_f_11': row[1][15:16]}

        elif row[1][0:3] == 'ADN': #VIS30 2-speed
            mmc = {'V302_f_01': row[1][0:3],
                   'V302_f_02': row[1][3:5],
                   'V302_f_03': row[1][5:6],
                   'V302_f_04': row[1][6:8],
                   'V302_f_05': row[1][8:9],
                   'V302_f_06': row[1][9:10],
                   'V302_f_07': row[1][10:11],
                   'V302_f_08': row[1][11:13],
                   'V302_f_09': row[1][13:14],
                   'V302_f_10': row[1][14:15],
                   'V302_f_11': row[1][15:16]}

        elif row[1][0:3] == 'ADT': #VIS40 2-speed
            mmc = {'V402_f_01': row[1][0:3],
                   'V402_f_02': row[1][3:5],
                   'V402_f_03': row[1][5:6],
                   'V402_f_04': row[1][6:8],
                   'V402_f_05': row[1][8:9],
                   'V402_f_06': row[1][9:10],
                   'V402_f_07': row[1][10:11],
                   'V402_f_08': row[1][11:13],
                   'V402_f_09': row[1][13:14],
                   'V402_f_10': row[1][14:15],
                   'V402_f_11': row[1][15:16]}

        elif row[1][0:3] == 'ADP': #VIS45 2-speed
            mmc = {'V452_f_01': row[1][0:3],
                   'V452_f_02': row[1][3:5],
                   'V452_f_03': row[1][5:6],
                   'V452_f_04': row[1][6:8],
                   'V452_f_05': row[1][8:9],
                   'V452_f_06': row[1][9:10],
                   'V452_f_07': row[1][10:11],
                   'V452_f_08': row[1][11:13],
                   'V452_f_09': row[1][13:14],
                   'V452_f_10': row[1][14:15],
                   'V452_f_11': row[1][15:16]}

        elif row[1][1:5] == 'HP31': #HP30
            mmc = {'HP31_f_01': row[1][0:1],
                   'HP31_f_02': row[1][1:4],
                   'HP31_f_03': row[1][4:6],
                   'HP31_f_04': row[1][6:8],
                   'HP31_f_05': row[1][8:10],
                   'HP31_f_06': row[1][10:12],
                   'HP31_f_07': row[1][12:14],
                   'HP31_f_08': row[1][14:16],
                   'HP31_f_09': row[1][16:17],
                   'HP31_f_10': row[1][17:18],
                   'HP31_f_11': row[1][18:19],
                   'HP31_f_12': row[1][19:20],
                   'HP31_f_13': row[1][20:21],
                   'HP31_f_14': row[1][21:23],
                   'HP31_f_15': row[1][23:25],
                   'HP31_f_16': row[1][25:27],
                   'HP31_f_17': row[1][27:29],
                   'HP31_f_18': row[1][29:30]}

        elif row[1][0:5] == 'SBX60': #SBX
            mmc = {'SBX60_f_01': row[1][0:5],
                   'SBX60_f_02': row[1][5:6],
                   'SBX60_f_03': row[1][6:7],
                   'SBX60_f_04': row[1][7:8],
                   'SBX60_f_05': row[1][8:9],
                   'SBX60_f_06': row[1][9:10],
                   'SBX60_f_07': row[1][10:12],
                   'SBX60_f_08': row[1][12:15],
                   'SBX60_f_09': row[1][15:18],
                   'SBX60_f_10': row[1][18:19],
                   'SBX60_f_11': row[1][19:21],
                   'SBX60_f_12': row[1][21:22],
                   'SBX60_f_13': row[1][22:24],
                   'SBX60_f_14': row[1][24:25],
                   'SBX60_f_15': row[1][25:26],
                   'SBX60_f_16': row[1][26:28],
                   'SBX60_f_17': row[1][28:30],
                   'SBX60_f_18': row[1][30:31]}


        done = np.empty([1, 5])
        for bom_position in xls.sheet_names:
            if bom_position not in excluded:
                df = pd.read_excel(xls, bom_position, dtype=str, header=1)
                conditions = [col for col in df.columns if col[0:8] in mmc.keys()]
                if conditions:
                    for c in conditions:
                        df = df[df[c].str.contains(mmc[c[0:8]])]
                    if df.shape[0] == 1:
                        if df.iloc[0, -1] == 'Active':
                            done = np.vstack((done, df.values[0][-5:]))
                    elif df.shape[0] == 2:
                        if df.iloc[0, -1] == 'Active':
                            #sheet_conflicts.append([row[0], row[1], bom_position, df.iloc[0, -5], df.iloc[0, -4], df.iloc[1, -5], df.iloc[1, -4]])
                            done = np.vstack((done, df.values[1][-5:]))
        bom = pd.DataFrame(done[1:,:-1], columns=['Part', 'Qty', 'Position', 'Description'], dtype=str)
        bom['Parent'] = row[0]
        bom['MMC'] = row[1]
        bom = bom[['Position', 'Part', 'Qty', 'Description', 'Parent', 'MMC']]
        bom.loc[bom['Qty'] == '0', 'Qty'] = ' '

        rows = bom.values.tolist()

        for r in rows:
            sheet.append(r)



    workbook.save(filename=target_file)
    workbook.close()

    return 1





if __name__ == "__main__":
    mmc_list = pd.read_excel(r'C:\Users\u331609\Desktop\S10 Eaton\Seba\mmc_s10_acka.xlsx')
    adr_constrain = pd.ExcelFile(r'C:\Users\u331609\Desktop\S10 Eaton\Seba\ADR_CaseTable_New_PNs_changed1.xlsx')
    acc_constrain = pd.ExcelFile(r'C:\Users\u331609\Desktop\S10 Eaton\Seba\ACC_CaseTable_New_PNs.xlsx')

    workbook = openpyxl.Workbook()
    sheet_header = workbook['Sheet']
    sheet_header.title = 'mmc'
    sheet_header.append(['Position', 'Part', 'Qty', 'Description', 'Parent', 'MMC'])
    workbook.save(filename=r'C:\Users\u331609\Desktop\S10 Eaton\mmc_test.xlsx')

    excluded = ['dr_c_900_INSTALLATION_DRAWING', 'dr_c_Service_information', 'dr_c_902', 'N01', 'N02', 'N03', 'N04', 'N05', 'N06']
    for i, row in mmc_list.iterrows():

        if row[1][0:3] == 'ADR':
            mmc = {'ADR_F_01': row[1][0:3],
               'ADR_F_02': row[1][3:4],
               'ADR_F_03': row[1][4:5],
               'ADR_F_04': row[1][5:6],
               'ADR_F_05': row[1][6:7],
               'ADR_F_06': row[1][7:9],
               'ADR_F_07': row[1][9:10],
               'ADR_F_08': row[1][10:11],
               'ADR_F_09': row[1][11:12],
               'ADR_F_10': row[1][12:14],
               'ADR_F_11': row[1][14:16],
               'ADR_F_12': row[1][16:18],
               'ADR_F_13': row[1][18:20],
               'ADR_F_14': row[1][20:21],
               'ADR_F_15': row[1][21:23],
               'ADR_F_16': row[1][23:24],
               'ADR_F_17': row[1][24:25],
               'ADR_F_18': row[1][25:26],
               'ADR_F_19': row[1][26:27],
               'ADR_F_20': row[1][27:29],
               'ADR_F_21': row[1][29:30],
               'ADR_F_22': row[1][30:31],
               'ADR_F_23': row[1][31:32]}

            xls = adr_constrain

        elif row[1][0:3] == 'ACC':
            mmc = {'ACC_f_01': row[1][0:3],
                   'ACC_f_02': row[1][3:4],
                   'ACC_f_03': row[1][4:5],
                   'ACC_f_04': row[1][5:6],
                   'ACC_f_05': row[1][6:8],
                   'ACC_f_06': row[1][8:9],
                   'ACC_f_07': row[1][9:10],
                   'ACC_f_08': row[1][10:11],
                   'ACC_f_09': row[1][11:13],
                   'ACC_f_10': row[1][13:15],
                   'ACC_f_11': row[1][15:17],
                   'ACC_f_12': row[1][17:18],
                   'ACC_f_13': row[1][18:20],
                   'ACC_f_14': row[1][20:21],
                   'ACC_f_15': row[1][21:22],
                   'ACC_f_16': row[1][22:23],
                   'ACC_f_17': row[1][23:24],
                   'ACC_f_18': row[1][24:26],
                   'ACC_f_19': row[1][26:27],
                   'ACC_f_20': row[1][27:28],
                   'ACC_f_21': row[1][28:29]}
            xls = acc_constrain
        else:
            print('error')


        done = np.empty([1, 5])
        for bom_position in xls.sheet_names:
            if bom_position not in excluded and 'KIT' not in bom_position:
                df = pd.read_excel(xls, bom_position, dtype= str, header=1)
                conditions = [col for col in df.columns if col[0:8] in mmc.keys()]
                for c in conditions:
                    df = df[df[c].str.contains(mmc[c[0:8]])]
                if df.shape[0] == 1:
                    done = np.vstack((done, df.values[0][-5:]))
                elif df.shape[0] == 2:
                    # print(f'some weird condition {row["Part"]}')
                    # print(bom_position)
                    # print(df.iloc[:, -5])
                    print(row["Part"])
                    workbook1 = openpyxl.load_workbook(filename=r'C:\Users\u331609\Desktop\S10 Eaton\conflicting conditions_acka.xlsx')
                    sheet1 = workbook1['Sheet1']
                    sheet1.append([row[0],row[1], bom_position, df.iloc[0, -5], df.iloc[0, -4], df.iloc[1, -5], df.iloc[1, -4]])
                    workbook1.save(filename=r'C:\Users\u331609\Desktop\S10 Eaton\conflicting conditions_acka.xlsx')
                    workbook1.close()
        # bom = pd.DataFrame(done[1:,:-1], columns=['Part', 'Qty', 'Position', 'Description'], dtype=str)
        # bom['Parent'] = row[0]
        # bom['MMC'] = row[1]
        # bom = bom[['Position', 'Part', 'Qty', 'Description', 'Parent', 'MMC']]
        # bom.loc[bom['Qty'] == '0', 'Qty'] = ' '
        #
        # rows = bom.values.tolist()
        # workbook = openpyxl.load_workbook(filename=r'C:\Users\u331609\Desktop\S10 Eaton\mmc_test.xlsx')
        # sheet = workbook['mmc']
        # for r in rows:
        #     sheet.append(r)
        # workbook.save(filename=r'C:\Users\u331609\Desktop\S10 Eaton\mmc_test.xlsx')


    adr_constrain.close()
    acc_constrain.close()
