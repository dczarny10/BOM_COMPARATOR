import openpyxl, re, os
import win32com.client as win32
import tempfile
import pandas as pd


def load_files(paths):
    pattern = "(\d+-?\w?\d+-?\d*P?C?H?M?)-(.+)"
    data = {}

    for file in paths:
        if file.lower().endswith("xls"):
            #excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel = win32.DispatchEx("Excel.Application")
            excel.DisplayAlerts = False
            excel.Visible = False
            wb = excel.Workbooks.Open(file)
            wb.SaveAs(tempfile.gettempdir()+"BOM_COMPARATOR_TEMP.xlsx", FileFormat=51)
            wb.Close()
            excel.Application.Quit()
            wb_SAP = openpyxl.load_workbook(tempfile.gettempdir()+"BOM_COMPARATOR_TEMP.xlsx")
            ws_SAP = wb_SAP.active
            ws_SAP.delete_cols(1, 1)
            ws_SAP.delete_rows(1, 3)
            ws_SAP.delete_cols(5, 10)
            for q in ws_SAP.iter_rows(values_only=True):
                product_name = str(q[0])+"_SAP"
                data.setdefault(product_name, [])
                index = -1
                for i, obj in enumerate(data[product_name]):
                    if obj[0] == q[1]:
                        index = i
                        break
                if index == -1:
                    data[product_name].append(tuple(map(str, (q[1], q[2], q[3].strip()))))
                else:
                    try:
                        data[product_name][index] = (data[product_name][index][0], str(int(data[product_name][index][1])+int(q[2])), data[product_name][index][2])
                    except ValueError:
                        print(product_name)
                        print(data[product_name][index][0])

            os.remove(tempfile.gettempdir()+"BOM_COMPARATOR_TEMP.xlsx")
            continue
        wb_PLM = openpyxl.load_workbook(file, data_only=True)
        ws_PLM = wb_PLM.worksheets[0]
        if ws_PLM.title == 'BOM Matrix':
            data_PLM = list(ws_PLM.iter_rows(values_only=True)) #create a list of all rows in excel file
            products = [i for i in data_PLM[0][3:]] #create products numbers list
            data_PLM = data_PLM[1:] #remove headlines in excel file
            for count, i in enumerate(products):
                i = i.strip()
                data.setdefault(i, [])
                for j in data_PLM:
                    if j[3+count] != "0": #if part belong to this product
                        r = re.findall(pattern, j[0].strip()) #find part number #python regular expression because part numbers are sadly not standarized
                        index = -1
                        for k, obj in enumerate(data[i]): #check if part is already under this product BOM
                            if obj[0] == r[0][0]:
                                index = k
                                break
                        if index == -1:
                            data[i].append((r[0][0], j[3 + count].split(".")[0], r[0][1].strip(), j[1] if j[1] is not None else ' '))
                        else:
                            try:
                                data[i][index] = (data[i][index][0], str(int(data[i][index][1]) + int(j[3 + count].split(".")[0])), data[i][index][2], str(data[i][index][3]) + ', ' + str(j[1].strip('0') if j[1] is not None else ' '))
                            except ValueError:
                                print(product_name)
                                print(data[product_name][index][0])
        elif ws_PLM.title == 'mmc':
            data_PLM = list(ws_PLM.iter_rows(values_only=True))
            data_PLM = data_PLM[1:]  # remove headlines in excel file
            products = set([i[4] for i in data_PLM if i[4] is not None])
            for count, i in enumerate(products):
                product_name = i.strip() + "_MMC"
                data.setdefault(product_name, [])
                for j in data_PLM:
                    if j[4].strip() == i.strip(): #if part belong to this product
                        r = j[1].strip()
                        index = -1
                        for k, obj in enumerate(data[product_name]): #check if part is already under this product BOM
                            if obj[0] == r:
                                index = k
                                break
                        if index == -1:
                            data[product_name].append((r, str(j[2]), j[3].strip(), j[0] if j[0] is not None else ' '))
                        else:
                            if data[product_name][index][1] == ' ' or j[2] == ' ':
                                data[product_name][index] = (data[product_name][index][0], ' ', data[product_name][index][2])
                            else:
                                try:
                                    data[product_name][index] = (data[product_name][index][0], str(int(data[product_name][index][1]) + int(j[2])), data[product_name][index][2], str(data[product_name][index][3]) + ', ' + str(j[0] if j[0] is not None else ' '))
                                except ValueError:
                                    print(product_name)
                                    print(data[product_name][index][0])
        else:
            data_PLM = list(ws_PLM.iter_rows(values_only=True))
            data_PLM = data_PLM[1:]  # remove headlines in excel file
            products = set([i[4] for i in data_PLM if i[4] is not None])
            for count, i in enumerate(products):
                product_name = i.strip() + "_PDF"
                data.setdefault(product_name, [])
                for j in data_PLM:
                    if j[4].strip() == i.strip(): #if part belong to this product
                        r = j[1].strip()
                        index = -1
                        for k, obj in enumerate(data[product_name]): #check if part is already under this product BOM
                            if obj[0] == r:
                                index = k
                                break
                        if index == -1:
                            data[product_name].append((r, str(j[2]), j[3].strip(), j[0] if j[0] is not None else ' '))
                        else:
                            try:
                                data[product_name][index] = (data[product_name][index][0], str(int(data[product_name][index][1]) + int(j[2])), data[product_name][index][2], str(data[product_name][index][3]) + ', ' + str(j[0] if j[0] is not None else ' '))
                            except ValueError:
                                print(product_name)
                                print(data[product_name][index][0])
    return data
